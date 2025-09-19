# project/services/export_services.py
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
import io


def export_to_excel(report_data, columns, filename='report.xlsx'):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Report'

    # Headers
    for col_num, column_title in enumerate(columns, 1):
        ws[f'{get_column_letter(col_num)}1'] = column_title

    # Data rows
    for row_num, row_data in enumerate(report_data, 2):
        for col_num, key in enumerate(columns, 1):
            val = row_data.get(key, '')
            # write raw values; openpyxl will format
            ws[f'{get_column_letter(col_num)}{row_num}'] = str(val) if val is not None else ''

    # Autosize (basic)
    for i, _ in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = 18

    # Prepare response
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def export_to_pdf(report_data, columns, filename='report.pdf'):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    style_sheet = getSampleStyleSheet()
    elements.append(Paragraph('Report', style_sheet['Title']))

    table_data = [columns] + [
        [str(row.get(col, '')) for col in columns] for row in report_data
    ]

    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
    ]))
    elements.append(table)
    doc.build(elements)

    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
