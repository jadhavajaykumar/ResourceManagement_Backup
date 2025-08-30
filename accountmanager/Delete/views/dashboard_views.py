# Views for Account Manager dashboard
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Sum
#from datetime import datetime
from accounts.utils import is_accountmanager
from employee.models import EmployeeProfile
import re
#from accountmanager.utils import is_accountmanager
from expenses.models import Expense, DailyAllowance, AdvanceRequest
from utils.currency import format_currency
from django.contrib import messages
from collections import defaultdict

from django.urls import reverse
from django.db.models import Q
from django.utils.dateparse import parse_date
import openpyxl
from openpyxl.utils import get_column_letter

from datetime import date




 # Adjust path as needed




from accountmanager.views.common import is_accountmanager  # ensure this role check exists



import io
import xlsxwriter
from django.http import HttpResponse

from django.contrib.auth import get_user_model


@login_required
@user_passes_test(is_accountmanager)
def export_settlement_summary_excel(request):
    User = get_user_model()

    expense_totals = Expense.objects.filter(
        status="Approved", reimbursed=False, forwarded_to_accountmanager=True
    ).values('employee', 'currency').annotate(total=Sum('amount'))

    da_totals = DailyAllowance.objects.filter(
        approved=True, reimbursed=False, forwarded_to_accountmanager=True
    ).values('employee', 'currency').annotate(total=Sum('da_amount'))

    # Group totals by employee and currency
    totals = {}
    for e in expense_totals:
        key = (e['employee'], e['currency'])
        totals[key] = totals.get(key, 0) + e['total']
    for d in da_totals:
        key = (d['employee'], d['currency'])
        totals[key] = totals.get(key, 0) + d['total']

    employee_map = {u.id: u.get_full_name() for u in User.objects.filter(id__in={k[0] for k in totals})}

    # Prepare Excel file
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    worksheet = workbook.add_worksheet('Settlement Summary')

    header_format = workbook.add_format({'bold': True, 'bg_color': '#D9E1F2'})
    currency_format = workbook.add_format({'num_format': '#,##0.00'})

    # Write headers
    headers = ['Employee Name', 'Currency', 'Total Unsettled Amount']
    for col_num, header in enumerate(headers):
        worksheet.write(0, col_num, header, header_format)

    # Write data rows
    for row_num, ((emp_id, currency), total) in enumerate(totals.items(), start=1):
        worksheet.write(row_num, 0, employee_map.get(emp_id, ''))
        worksheet.write(row_num, 1, currency)
        worksheet.write_number(row_num, 2, float(total), currency_format)

    workbook.close()
    output.seek(0)

    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="settlement_summary.xlsx"'
    return response


@login_required
@user_passes_test(is_accountmanager)
def dashboard_view(request):
    profile = EmployeeProfile.objects.get(user=request.user)
    return render(request, 'accountmanager/dashboard.html', {'profile': profile})





# accountmanager/views/dashboard_views.py




@login_required
@user_passes_test(is_accountmanager)
def reimbursement_dashboard(request):
    # Tab 1: Pending Reimbursement – Manager Approved
    expenses = Expense.objects.filter(
        status="Forwarded to Account Manager",
        reimbursed=False,
        forwarded_to_accountmanager=True
    ).select_related('employee__user', 'project')

    allowances = DailyAllowance.objects.filter(
        approved=True,
        reimbursed=False,
        forwarded_to_accountmanager=True
    ).select_related('employee__user', 'project')

    reimbursement_entries = []

    # Add expenses
    for expense in expenses:
        reimbursement_entries.append({
            "type": "Expense",
            "employee": expense.employee,
            "employee_name": expense.employee.user.get_full_name(),
            "project": expense.project,
            "amount": format_currency(expense.amount, getattr(expense, 'currency', 'INR')),
            "date": expense.date,
            "expense_type": expense.new_expense_type,
            "id": expense.id,
        })

    # Add allowances
    for da in allowances:
        reimbursement_entries.append({
            "type": "DA",
            "employee": da.employee,
            "employee_name": da.employee.user.get_full_name(),
            "project": da.project,
            "amount": format_currency(da.da_amount, da.currency),
            "date": da.date,
            "expense_type": "DA",
            "id": da.id,
        })

    # Tab 2: Grouped summary by employee
    employee_summaries = {}
    for entry in reimbursement_entries:
        emp = entry["employee"]
        emp_id = emp.id
        amt_str = entry["amount"]
        symbol = amt_str[0]
        amount = float(amt_str[1:].replace(",", "").strip())

        if emp_id not in employee_summaries:
            employee_summaries[emp_id] = {
                "employee": emp,
                "total_amount": 0.0,
                "currency_symbol": symbol,
            }
        employee_summaries[emp_id]["total_amount"] += amount

    grouped_summary = []
    for summary in employee_summaries.values():
        total_amt = summary["total_amount"]
        summary["total_formatted"] = f"{summary['currency_symbol']} {total_amt:,.2f}"
        grouped_summary.append(summary)

    # Tab 3: Settled history
    settled_expenses = Expense.objects.filter(
        status="Approved", reimbursed=True
    ).select_related('employee__user', 'project')

    settled_das = DailyAllowance.objects.filter(
        approved=True, reimbursed=True
    ).select_related('employee__user', 'project')

    settled_employee_ids = set(e.employee.id for e in settled_expenses) | set(d.employee.id for d in settled_das)
    employee_map = {
        emp.id: emp for emp in EmployeeProfile.objects.select_related('user').filter(id__in=settled_employee_ids)
    }

    settled_data = defaultdict(list)
    for e in settled_expenses:
        settled_data[e.employee.id].append({
            "type": "Expense",
            "employee": e.employee,
            "project": e.project,
            "amount": format_currency(e.amount, getattr(e, 'currency', 'INR')),
            "date": e.date,
        })

    for d in settled_das:
        settled_data[d.employee.id].append({
            "type": "DA",
            "employee": d.employee,
            "project": d.project,
            "amount": format_currency(d.da_amount, d.currency),
            "date": d.date,
        })

    settled_merged = {}
    for emp_id, entries in settled_data.items():
        total = 0.0
        currency_symbol = "₹"
        for entry in entries:
            amt_str = entry["amount"]
            if amt_str and len(amt_str) > 1:
                currency_symbol = amt_str[0]
                try:
                    amt = float(amt_str[1:].replace(",", ""))
                    total += amt
                except Exception:
                    pass
        settled_merged[emp_id] = {
            "employee": employee_map.get(emp_id),
            "total": f"{currency_symbol}{total:,.2f}",
            "entries": entries,
        }

    # Tab 4: Approved advances (ready for settlement)
    approved_advances = AdvanceRequest.objects.filter(
        approved_by_manager=True,
        approved_by_accountant=True,
        settled_by_account_manager=False
    ).select_related('employee__user', 'project')

    return render(request, 'accountmanager/reimbursement_dashboard.html', {
        'reimbursement_entries': reimbursement_entries,
        'employee_summaries': grouped_summary,
        'settled_data': settled_data,
        'settled_merged': settled_merged,
        'employee_map': employee_map,
        'tab_data': {
            'advances': approved_advances
        },
        'active_tab': request.GET.get("tab", "tab1"),
    })










@login_required
@user_passes_test(is_accountmanager)
def settlement_summary(request):
    from django.contrib.auth import get_user_model
    User = get_user_model()

    expense_totals = Expense.objects.filter(
        status="Approved", reimbursed=False, forwarded_to_AccountManager=True
    ).values('employee', 'currency').annotate(total=Sum('amount'))

    da_totals = DailyAllowance.objects.filter(
        approved=True, reimbursed=False, forwarded_to_AccountManager=True
    ).values('employee', 'currency').annotate(total=Sum('da_amount'))

    # Group by employee and currency
    totals = {}
    for e in expense_totals:
        key = (e['employee'], e['currency'])
        totals[key] = totals.get(key, 0) + e['total']
    for d in da_totals:
        key = (d['employee'], d['currency'])
        totals[key] = totals.get(key, 0) + d['total']

    employee_map = {u.id: u.get_full_name() for u in User.objects.filter(id__in={k[0] for k in totals})}
    final_data = [
        {
            'employee_id': emp_id,
            'employee': employee_map.get(emp_id, ''),
            'total': total,
            'currency': currency
        }
        for (emp_id, currency), total in totals.items()
    ]

    return render(request, 'accountmanager/settlement_summary.html', {
        'totals': final_data,
        'format_currency': format_currency,
    })


@login_required
@user_passes_test(is_accountmanager)
def settle_employee(request, employee_id):
    if request.method == "POST":
        Expense.objects.filter(employee_id=employee_id, status="Approved", reimbursed=False, forwarded_to_AccountManager=True).update(
            reimbursed=True
        )
        DailyAllowance.objects.filter(employee_id=employee_id, approved=True, reimbursed=False, forwarded_to_AccountManager=True).update(
            reimbursed=True
        )
        return redirect('accountmanager:settlement-summary')
    return redirect('accountmanager:settlement-summary')


@login_required
@user_passes_test(is_accountmanager)
def settle_selected(request):
    if request.method == "POST":
        employee_id = request.POST.get("settle_employee")
        reimbursement_date = request.POST.get(f"settlement_date_{employee_id}")

        if not employee_id:
            messages.error(request, "No employee selected for reimbursement.")
            return redirect(f"{reverse('accountmanager:reimbursement_dashboard')}?tab=tab2")

        if not reimbursement_date:
            messages.error(request, "Please enter a settlement date.")
            return redirect(f"{reverse('accountmanager:reimbursement_dashboard')}?tab=tab2")

        # Mark expenses & DA as reimbursed
        Expense.objects.filter(
            employee_id=employee_id,
            status="Forwarded to Account Manager",
            reimbursed=False,
            forwarded_to_accountmanager=True
        ).update(
            reimbursed=True,
            status="Approved",
            settlement_date=reimbursement_date
        )

        DailyAllowance.objects.filter(
            employee_id=employee_id,
            approved=True,
            reimbursed=False,
            forwarded_to_accountmanager=True
        ).update(
            reimbursed=True,
            settlement_date=reimbursement_date
        )

        messages.success(request, f"Reimbursement settled for employee {employee_id}.")

        # Redirect back to Summary by Employee tab
        return redirect(f"{reverse('accountmanager:reimbursement_dashboard')}?tab=tab2")



@login_required
def export_tab1(request):
    headers = ['Employee', 'Date', 'Amount', 'Type']
    rows = [
        ["Ajay Kumar", "2025-06-20", "₹100", "Expense"],
        # Populate from tab 1 actual data
    ]
    return generate_excel_response("Tab1_Report", headers, rows)

@login_required
def export_tab2(request):
    headers = ['Employee', 'Total Pending Amount', 'Entry Count']
    rows = [
        ["Ajay Kumar", "₹600", "3"],
        # Populate from tab 2 actual data
    ]
    return generate_excel_response("Tab2_PendingSummary", headers, rows)

@login_required
def export_tab3(request):
    view_type = request.GET.get("view", "detailed")

    settled_expenses = Expense.objects.select_related('employee__user', 'project').filter(status="Approved", reimbursed=True)
    settled_das = DailyAllowance.objects.select_related('employee__user', 'project').filter(approved=True, reimbursed=True)

    settled_data = defaultdict(list)
    employee_ids = set()

    for e in settled_expenses:
        emp = e.employee
        employee_ids.add(emp.id)
        settled_data[emp.id].append({
            "type": "Expense",
            "employee": emp,
            "project": e.project,
            "amount": format_currency(e.amount, getattr(e, 'currency', 'INR')),
            "date": e.date,
        })

    for d in settled_das:
        emp = d.employee
        employee_ids.add(emp.id)
        settled_data[emp.id].append({
            "type": "DA",
            "employee": emp,
            "project": d.project,
            "amount": format_currency(d.da_amount, d.currency),
            "date": d.date,
        })

    employee_map = {
        emp.id: emp for emp in EmployeeProfile.objects.select_related('user').filter(id__in=employee_ids)
    }

    if view_type == "detailed":
        headers = ['Employee', 'Date', 'Type', 'Amount', 'Project']
        rows = []
        for emp_id, entries in settled_data.items():
            emp = employee_map.get(emp_id)
            if not emp: continue
            for entry in entries:
                rows.append([
                    emp.user.get_full_name(),
                    entry['date'].strftime("%Y-%m-%d"),
                    entry['type'],
                    entry['amount'],
                    entry['project'].name if entry['project'] else "-"
                ])
        return generate_excel_response("Tab3_Detailed", headers, rows)

    else:  # Merged
        headers = ['Employee', 'Total Amount', 'Entry Count']
        rows = []
        for emp_id, entries in settled_data.items():
            emp = employee_map.get(emp_id)
            if not emp: continue
            total = 0
            currency = '₹'
            for entry in entries:
                amt = float(entry['amount'][1:].replace(",", ""))
                total += amt
                currency = entry['amount'][0]
            rows.append([
                emp.user.get_full_name(),
                f"{currency}{total:,.2f}",
                len(entries)
            ])
        return generate_excel_response("Tab3_Merged", headers, rows)




def generate_excel_response(filename_prefix, headers, rows):
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    worksheet = workbook.add_worksheet()

    # Write headers
    for col_num, header in enumerate(headers):
        worksheet.write(0, col_num, header)

    # Write data rows
    for row_num, row_data in enumerate(rows, start=1):
        for col_num, cell_value in enumerate(row_data):
            worksheet.write(row_num, col_num, cell_value)

    workbook.close()
    output.seek(0)

    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"{filename_prefix}_{date.today()}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
