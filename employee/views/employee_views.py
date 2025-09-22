# employee/views/employee_views.py
from django.views.decorators.http import require_POST
from django.core.files.storage import FileSystemStorage

import openpyxl
from django.http import HttpResponse
from django.contrib.admin.views.decorators import staff_member_required
from django.utils import timezone
from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from accounts.models import CustomUser
from employee.models import EmployeeProfile  # generate_employee_id removed
from employee.forms import EmployeeProfileForm
from django.db import transaction
from django.contrib.auth.hashers import make_password
from django.contrib.auth.models import Group
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Q
from datetime import datetime

# helper: only HR/Admin
def is_hr_or_admin(user):
    return user.is_superuser or user.groups.filter(name='HR').exists() or user.has_perm('timesheet.can_approve')


@login_required
def employee_list(request):
    """
    Server-side filtered, paginated Employee Directory view for HR/staff.
    Provides context keys expected by the template:
      - employees (page object .object_list)
      - page_obj, paginator, is_paginated
      - filters (dict of current filter values)
      - request_get (request.GET)
      - can_edit (bool based on is_hr_or_admin)
      - page_sizes (list)
    """
    qs = EmployeeProfile.objects.select_related('user', 'reporting_manager').all().order_by('user__first_name', 'user__last_name')

    # Parse filters from GET
    filters = {
        'q': request.GET.get('q', '').strip(),
        'name': request.GET.get('name', '').strip(),
        'employee_id': request.GET.get('employee_id', '').strip(),
        'role': request.GET.get('role', '').strip(),
        'department': request.GET.get('department', '').strip(),
        'employment_type': request.GET.get('employment_type', '').strip(),
        'contact': request.GET.get('contact', '').strip(),
        'manager': request.GET.get('manager', '').strip(),
        'address': request.GET.get('address', '').strip(),
        'page_size': int(request.GET.get('page_size') or 25),
    }

    # Global query search across common fields
    if filters['q']:
        q = filters['q']
        qs = qs.filter(
            Q(user__first_name__icontains=q) |
            Q(user__last_name__icontains=q) |
            Q(user__email__icontains=q) |
            Q(employee_id__icontains=q) |
            Q(department__icontains=q) |
            Q(address__icontains=q)
        )

    # Field-level filters
    if filters['name']:
        n = filters['name']
        qs = qs.filter(
            Q(user__first_name__icontains=n) |
            Q(user__last_name__icontains=n) |
            Q(user__email__icontains=n)
        )
    if filters['employee_id']:
        qs = qs.filter(employee_id__icontains=filters['employee_id'])
    if filters['role']:
        qs = qs.filter(role__icontains=filters['role'])
    if filters['department']:
        qs = qs.filter(department__icontains=filters['department'])
    if filters['employment_type']:
        qs = qs.filter(employment_type__icontains=filters['employment_type'])
    if filters['contact']:
        qs = qs.filter(contact_number__icontains=filters['contact'])
    if filters['manager']:
        m = filters['manager']
        qs = qs.filter(
            Q(reporting_manager__first_name__icontains=m) |
            Q(reporting_manager__last_name__icontains=m) |
            Q(reporting_manager__email__icontains=m)
        )
    if filters['address']:
        qs = qs.filter(address__icontains=filters['address'])

    # Pagination
    page_size = filters['page_size'] or 25
    page = request.GET.get('page') or 1
    paginator = Paginator(qs, page_size)
    employees_page = paginator.get_page(page)

    context = {
        'employees': employees_page.object_list,
        'page_obj': employees_page,
        'paginator': paginator,
        'is_paginated': employees_page.has_other_pages(),
        'filters': filters,
        'request_get': request.GET,
        'can_edit': is_hr_or_admin(request.user),
        'page_sizes': [10, 25, 50, 100],  # avoid template parsing issues when using a tuple literal
    }

    return render(request, 'employee/employee_list.html', context)


@login_required
@user_passes_test(is_hr_or_admin)
def employee_create(request, pk=None):
    """
    Create or edit an EmployeeProfile. If pk provided -> edit.
    NOTE: employee_id must be provided via form (no auto-generation).
    """
    instance = None
    if pk:
        instance = get_object_or_404(EmployeeProfile, pk=pk)

    if request.method == 'POST':
        form = EmployeeProfileForm(request.POST, instance=instance)
        if form.is_valid():
            # DO NOT auto-generate employee_id anymore. The form enforces presence/uniqueness.
            obj = form.save(commit=False)
            obj.save()
            messages.success(request, "Employee profile saved.")
            return redirect('employee:employee-list')
        else:
            messages.error(request, "Please correct the highlighted errors.")
    else:
        form = EmployeeProfileForm(instance=instance)

    # Useful context for template: existing users (to create user from admin separately)
    users_count = CustomUser.objects.count()
    return render(request, 'employee/employee_form.html', {
        'form': form,
        'editing': bool(instance),
        'users_count': users_count,
    })


@staff_member_required
def export_employee_template(request):
    """
    Export an Excel file with headers for EmployeeProfile import.
    If employees exist, include them as sample rows.
    """
    headers = [
        "first_name", "last_name", "email", "role", "career_start_date",
        "probotix_joining_date", "date_of_birth", "contact_number", "address",
        "emergency_contact_name", "emergency_contact_relation", "emergency_contact_number",
        "employee_id", "department", "reporting_manager_email",
        "employment_type", "pan_aadhar_ssn", "bank_account_number",
        "bank_ifsc_code", "epf_number", "grace_period_days"
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employees"

    # write headers
    ws.append(headers)

    # add sample data (existing employees, limited to 20 rows for safety)
    for emp in EmployeeProfile.objects.select_related("user", "reporting_manager")[:20]:
        ws.append([
            emp.user.first_name,
            emp.user.last_name,
            emp.user.email,
            emp.role,
            emp.career_start_date.isoformat() if emp.career_start_date else "",
            emp.probotix_joining_date.isoformat() if emp.probotix_joining_date else "",
            emp.date_of_birth.isoformat() if emp.date_of_birth else "",
            emp.contact_number,
            emp.address,
            emp.emergency_contact_name,
            emp.emergency_contact_relation,
            emp.emergency_contact_number,
            emp.employee_id,
            emp.department,
            emp.reporting_manager.email if emp.reporting_manager else "",
            emp.employment_type,
            emp.pan_aadhar_ssn,
            emp.bank_account_number,
            emp.bank_ifsc_code,
            emp.epf_number,
            emp.grace_period_days,
        ])

    # response
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    filename = f"employee_import_template_{timezone.now().date()}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@staff_member_required
def import_employees(request):
    """
    Import employees from uploaded Excel file. Expects the template exported earlier.
    Validates email existence. If email missing or invalid, row is skipped and error reported.
    IMPORTANT: employee_id must be present in the sheet. Rows missing employee_id are marked as errors.
    """
    if request.method == "POST" and request.FILES.get("file"):
        file = request.FILES["file"]
        errors = []
        created = 0
        updated = 0

        try:
            wb = openpyxl.load_workbook(file)
        except Exception as e:
            messages.error(request, f"Could not read Excel file: {e}")
            return redirect(request.path)

        ws = wb.active
        headers = [ (cell.value or "").strip() for cell in ws[1] ]

        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # skip fully empty rows
            if not any(row):
                continue
            data = dict(zip(headers, row or []))

            # email required to link to CustomUser
            email_raw = (data.get("email") or "")
            email = str(email_raw).strip() if email_raw is not None else ""
            if not email:
                errors.append(f"Row {idx}: email is required")
                continue

            # employee_id MUST be provided (no auto-generation)
            raw_empid = data.get("employee_id")
            emp_id = str(raw_empid).strip() if raw_empid is not None else ""
            if not emp_id:
                errors.append(f"Row {idx}: employee_id is required (please fill in EMP ID in the spreadsheet).")
                continue

            # prefer existing CustomUser; if not exist create a minimal CustomUser
            try:
                user = CustomUser.objects.filter(email__iexact=email).first()
                if not user:
                    # create a minimal CustomUser to link the profile to
                    # Use email as username to avoid NULL username error
                    user = CustomUser.objects.create(
                        email=email,
                        username=email,
                        first_name=(data.get("first_name") or "")[:150],
                        last_name=(data.get("last_name") or "")[:150],
                    )
                else:
                    # update names if changed
                    fn = (data.get("first_name") or "")[:150]
                    ln = (data.get("last_name") or "")[:150]
                    changed = False
                    if user.first_name != fn:
                        user.first_name = fn
                        changed = True
                    if user.last_name != ln:
                        user.last_name = ln
                        changed = True
                    if user.username != email:
                        user.username = email
                        changed = True
                    if changed:
                        user.save()
            except Exception as e:
                errors.append(f"Row {idx}: error creating/looking up user: {e}")
                continue

            # resolve manager (by email)
            mgr_email = (data.get("reporting_manager_email") or "").strip()
            manager_user = CustomUser.objects.filter(email__iexact=mgr_email).first() if mgr_email else None

            # create/update EmployeeProfile
            try:
                # Check collision: if employee_id exists and belongs to a different user -> error
                existing = EmployeeProfile.objects.filter(employee_id=emp_id).first()
                if existing and existing.user_id != user.id:
                    errors.append(f"Row {idx}: employee_id '{emp_id}' already assigned to another user ({existing.user.email}).")
                    continue

                emp_defaults = {
                    "role": (data.get("role") or "Employee"),
                    "employee_id": emp_id,
                }
                emp, created_flag = EmployeeProfile.objects.get_or_create(user=user, defaults=emp_defaults)

                # If profile existed but employee_id empty or different, set/fix if it belongs to same user
                if not created_flag:
                    if (not emp.employee_id) or (emp.employee_id != emp_id):
                        # This is same user, allow setting/updating employee_id
                        emp.employee_id = emp_id

                # update fields - parse dates safely (if strings present keep unchanged)
                def parse_date_safe(val):
                    if not val:
                        return None
                    if isinstance(val, (str,)):
                        try:
                            return datetime.fromisoformat(val).date()
                        except Exception:
                            # try common formats? currently return None (skip)
                            return None
                    return val

                # set fields only if present in the sheet, else keep existing
                if data.get("career_start_date"):
                    parsed = parse_date_safe(data.get("career_start_date"))
                    if parsed:
                        emp.career_start_date = parsed
                if data.get("probotix_joining_date"):
                    parsed = parse_date_safe(data.get("probotix_joining_date"))
                    if parsed:
                        emp.probotix_joining_date = parsed
                if data.get("date_of_birth"):
                    parsed = parse_date_safe(data.get("date_of_birth"))
                    if parsed:
                        emp.date_of_birth = parsed

                emp.contact_number = data.get("contact_number") or emp.contact_number
                emp.address = data.get("address") or emp.address
                emp.emergency_contact_name = data.get("emergency_contact_name") or emp.emergency_contact_name
                emp.emergency_contact_relation = data.get("emergency_contact_relation") or emp.emergency_contact_relation
                emp.emergency_contact_number = data.get("emergency_contact_number") or emp.emergency_contact_number
                emp.department = data.get("department") or emp.department
                emp.reporting_manager = manager_user or emp.reporting_manager
                emp.employment_type = data.get("employment_type") or emp.employment_type
                emp.pan_aadhar_ssn = data.get("pan_aadhar_ssn") or emp.pan_aadhar_ssn
                emp.bank_account_number = data.get("bank_account_number") or emp.bank_account_number
                emp.bank_ifsc_code = data.get("bank_ifsc_code") or emp.bank_ifsc_code
                emp.epf_number = data.get("epf_number") or emp.epf_number
                emp.grace_period_days = data.get("grace_period_days") or emp.grace_period_days

                emp.save()

                if created_flag:
                    created += 1
                else:
                    updated += 1

            except Exception as e:
                errors.append(f"Row {idx}: error saving EmployeeProfile: {e}")
                continue

        return render(request, "employee/import_result.html", {
            "created": created,
            "updated": updated,
            "errors": errors,
        })

    return render(request, "employee/import_form.html")


# ---- bulk user import/export (append to employee/views/employee_views.py) ----

@staff_member_required
def export_user_template(request):
    """
    Export Excel template for bulk-creating users (admin).
    Columns:
      username, email, first_name, last_name, temp_password, is_staff, is_superuser, groups
    """
    headers = [
        "username", "email", "first_name", "last_name",
        "temp_password", "is_staff", "is_superuser", "groups"
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Users"
    ws.append(headers)

    # sample row (helpful)
    ws.append([
        "ajay.k@example.com", "ajay.k@example.com", "Ajay", "Kumar", "TempP@ss123", "True", "False", "HR"
    ])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    filename = f"user_import_template_{timezone.now().date()}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@staff_member_required
def import_users_accounts(request):
    """
    Admin view to import users from Excel.
    Expected headers: username, email, first_name, last_name, temp_password, is_staff, is_superuser, groups
    If username blank, 'email' will be used as username.
    Creates CustomUser records only (no EmployeeProfile).
    Robust: validates email, skips empty rows, collects row-level errors rather than crashing.
    """
    if request.method == "POST" and request.FILES.get("file"):
        file = request.FILES["file"]
        errors = []
        created = 0
        updated = 0

        try:
            wb = openpyxl.load_workbook(file)
        except Exception as e:
            messages.error(request, f"Could not read Excel file: {e}")
            return redirect(request.path)

        ws = wb.active
        headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]

        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # skip fully empty rows
            if not any(row):
                continue
            data = dict(zip(headers, row or []))

            row_errors = []
            # email is mandatory
            email_raw = (data.get("email") or "")
            email = str(email_raw).strip() if email_raw is not None else ""
            if not email:
                row_errors.append("email is required")
                errors.append(f"Row {idx}: {', '.join(row_errors)}")
                continue

            username_raw = data.get("username") or ""
            username = str(username_raw).strip() or email

            # optional fields
            first_name = (data.get("first_name") or "")[:150]
            last_name  = (data.get("last_name") or "")[:150]
            temp_password = data.get("temp_password") or ""
            is_staff = str(data.get("is_staff") or "False").strip().lower() in ("1","true","yes","y")
            is_superuser = str(data.get("is_superuser") or "False").strip().lower() in ("1","true","yes","y")
            groups_raw = (data.get("groups") or "") or ""
            group_names = [g.strip() for g in groups_raw.split(",") if g.strip()]

            try:
                with transaction.atomic():
                    # Ensure no attempt to write NULL username — fallback to email already handled
                    user, created_flag = CustomUser.objects.get_or_create(
                        email=email,
                        defaults={
                            "username": username,
                            "first_name": first_name,
                            "last_name": last_name,
                            "is_staff": is_staff,
                            "is_superuser": is_superuser,
                        }
                    )

                    modified = False
                    if not created_flag:
                        # update changed fields (safe)
                        if user.username != username:
                            user.username = username
                            modified = True
                        if user.first_name != first_name:
                            user.first_name = first_name
                            modified = True
                        if user.last_name != last_name:
                            user.last_name = last_name
                            modified = True
                        if user.is_staff != is_staff:
                            user.is_staff = is_staff
                            modified = True
                        if user.is_superuser != is_superuser:
                            user.is_superuser = is_superuser
                            modified = True

                    if temp_password:
                        user.password = make_password(temp_password)
                        modified = True

                    # save once
                    user.save()

                    # groups
                    if group_names:
                        groups_objs = []
                        for gname in group_names:
                            grp, _ = Group.objects.get_or_create(name=gname)
                            groups_objs.append(grp)
                        user.groups.set(groups_objs)

                    if created_flag:
                        created += 1
                    elif modified:
                        updated += 1

            except Exception as e:
                # Capture the row error and continue
                errors.append(f"Row {idx}: {str(e)}")

        return render(request, "employee/import_users_result.html", {
            "created": created,
            "updated": updated,
            "errors": errors,
        })

    return render(request, "employee/import_users_form.html")
